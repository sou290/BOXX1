import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

st.title("📦 Remplissage automatique de fiche BOX")


# ✅ Uploader maintenant accepte les fichiers .xls au lieu de .xlsb
uploaded_source = st.file_uploader("🗂️ Uploadez le fichier d'extraction (Excel)", type="xls")
=======
# Upload des fichiers.
uploaded_source = st.file_uploader("🗂️ Uploadez le fichier d'extraction (Excel)", type="xlsb")
uploaded_template = st.file_uploader("📄 Uploadez la fiche BOX vierge (template Excel)", type="xlsx")

# Entrée du code produit
code_produit = st.text_input("🔍 Entrez le code produit à filtrer")

if uploaded_source and uploaded_template and code_produit:
    try:

        # Lire temporairement toutes les lignes sans en-têtes
        temp_df = pd.read_excel(uploaded_source, header=None)

        # Identifier la ligne contenant les en-têtes (ex: "Product code")
=======
        # Lire toutes les lignes temporairement sans en-têtes
        temp_df = pd.read_excel(uploaded_source, header=None)

        # Chercher la ligne qui contient les vraies colonnes (ex: "Product code")

        row_index = None
        for i, row in temp_df.iterrows():
            if "Product code" in row.values:
                row_index = i
                break

        if row_index is None:
            st.error("❌ Impossible de localiser la ligne contenant les en-têtes (ex : 'Product code').")
            st.stop()

        # Lire à partir de la bonne ligne détectée
        df_source = pd.read_excel(uploaded_source, skiprows=row_index)

        # Filtrer selon le code produit entré
        df_filtré = df_source[df_source['Product code'].astype(str) == code_produit]

        if df_filtré.empty:
            st.error("❌ Code produit non trouvé dans l'extraction.")
        else:
            ligne = df_filtré.iloc[0]
            wb = load_workbook(uploaded_template)
            ws = wb.worksheets[0]

            mapping = {
                'Master barcode': 'C5',
                'Merchandise structure': 'C6',
                'Product code': 'E5',
                'Supplier Reference': 'E6',
                'Supplier Name': 'E7',
                'Origin': 'C9',
                'Brand name': 'C10',
                'Product short description': 'G6',
                'Description': 'G7',
                'Master Height  cm ': 'D16',
                'Master Width  cm ': 'D17',
                'Master Length  cm ': 'D18',
                'Palett Height  cm ': 'H16',
                'Master Weight  kg ': 'D19',
                'Palett Weight  kg ': 'H19',
                'Level / palet': 'E24',
                'CTN / level': 'E23'
            }

            convertir_en_mm = ['D17', 'D18', 'H16', 'D16']

            for colonne, cellule in mapping.items():
                if colonne in df_filtré.columns:
                    valeur = ligne[colonne]
                    if pd.isna(valeur):
                        valeur = ""
                    else:
                        if cellule in convertir_en_mm:
                            try:
                                valeur = int(float(str(valeur).replace(',', '.').strip()) * 10)
                            except:
                                valeur = ""
                    ws[cellule] = valeur
            output = io.BytesIO()
            wb.save(output)
            st.success("✅ Fiche remplie avec succès !")
            st.download_button("⬇️ Télécharger la fiche BOX remplie", data=output.getvalue(),
                               file_name=f"fiche_box_{code_produit}.xlsx")

    except Exception as e:
        st.error(f"💥 Une erreur s'est produite : {e}")
